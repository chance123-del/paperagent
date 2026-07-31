from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook

from paperformat_agent.annotations import load_annotations
from paperformat_agent.hybrid_insert import build_block


class AnnotationTests(unittest.TestCase):
    def _workbook(self, path: Path) -> None:
        workbook = Workbook()
        figures = workbook.active
        figures.title = "Figures"
        headers = ["asset_id", "caption_body", "caption_as_provided", "prefix_policy", "note", "source", "alt_text"]
        figures.append(headers)
        figures.append(["Fig1", "Architecture overview", "Fig. 1 Architecture overview", "template", "", "", ""])
        tables = workbook.create_sheet("Tables")
        tables.append(headers)
        tables.append(["Table1", "Dataset summary", "", "template", "", "", ""])
        links = workbook.create_sheet("Links")
        links.append(["asset_id", "url_or_doi", "link_text"])
        links.append(["Fig1", "10.1000/example", "Source"])
        workbook.save(path)

    def test_workbook_uses_caption_body_and_normalises_doi(self) -> None:
        with TemporaryDirectory() as directory:
            workbook = Path(directory) / "annotations.xlsx"
            self._workbook(workbook)
            result = load_annotations(str(workbook), Path(directory))
            self.assertEqual(result.figures["fig1"], "Architecture overview")
            self.assertEqual(result.tables["table1"], "Dataset summary")
            self.assertEqual(result.links["fig1"], ("https://doi.org/10.1000/example", "Source"))
            self.assertFalse(result.warnings)

    def test_caption_link_is_user_provided_and_escaped(self) -> None:
        with TemporaryDirectory() as directory:
            image = Path(directory) / "Fig1.png"
            image.write_bytes(b"not a real image")
            block = build_block("Figure", "", str(image), "Caption", Path(directory), caption_link=("https://x.test/a_b", "See source"))
            self.assertIn(r"\href{https://x.test/a\_b}{See source}", block)


if __name__ == "__main__":
    unittest.main()
