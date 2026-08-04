from __future__ import annotations

import csv
from pathlib import Path
import re
import shutil
from urllib.parse import urlparse


def _escape(value: str) -> str:
    return "".join({"&": r"\&", "%": r"\%", "_": r"\_", "#": r"\#"}.get(char, char) for char in value)


def _read_table(text: str, upload: str | None) -> list[list[str]]:
    if upload:
        path = Path(upload)
        if path.suffix.lower() in {".csv", ".tsv"}:
            delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
            with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
                return [row for row in csv.reader(handle, delimiter=delimiter) if any(cell.strip() for cell in row)]
        if path.suffix.lower() in {".xlsx", ".xls"}:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            return [["" if cell is None else str(cell) for cell in row] for row in workbook.active.iter_rows(values_only=True) if any(cell is not None for cell in row)]
    rows = []
    for line in text.splitlines():
        if "|" in line and not set(line.replace("|", "")) <= {"-", ":", " "}:
            rows.append([cell.strip() for cell in line.strip().strip("|").split("|")])
    return rows


def _insertion_policy(rules: dict | None) -> dict:
    base = {
        "figure_width": "0.85\\linewidth",
        "figure_alignment": "\\centering",
        "figure_float": "H",
        "table_alignment": "\\centering",
        "table_float": "H",
        "table_columns": "l",
        "hyperlink_command": "href",
    }
    if rules:
        base.update(rules.get("insertion_policy", {}))
    return base


def _caption_label(kind: str, caption: str, rules: dict | None) -> str:
    if caption.strip():
        return _escape(caption.strip())
    caption_prefixes = (rules or {}).get("caption_prefixes", {})
    return _escape(caption_prefixes.get(kind, kind.title()))


def build_block(
    kind: str,
    content: str | None,
    upload: str | None,
    caption: str | None,
    project_dir: Path,
    rules: dict | None = None,
    caption_link: tuple[str, str] | None = None,
) -> str:
    content = content or ""
    caption = caption or ""
    policy = _insertion_policy(rules)
    if kind == "Hyperlink":
        if not upload or not content.strip():
            raise ValueError("Enter link text and a URL for a hyperlink insertion.")
        parsed = urlparse(upload.strip())
        if parsed.scheme.lower() not in {"http", "https", "mailto"} or any(char in upload for char in "{}\r\n"):
            raise ValueError("Use a valid http, https, or mailto URL.")
        command = str(policy.get("hyperlink_command", "href")).strip() or "href"
        return chr(92) + command + "{" + upload.strip() + "}{" + _escape(content.strip()) + "}"
    if kind == "Formula":
        formula = content.strip()
        if not formula:
            raise ValueError("Enter a LaTeX formula before inserting.")
        forbidden = (r"\documentclass", r"\usepackage", r"\begin{document}", r"\end{document}", r"\input", r"\include", r"\write18")
        if any(token in formula.lower() for token in forbidden):
            raise ValueError("The formula contains a document-level or unsafe LaTeX command.")
        return "\\begin{equation}\n" + formula + "\n\\end{equation}"
    if kind == "Figure":
        if not upload:
            raise ValueError("Upload an image for a figure insertion.")
        source = Path(upload)
        assets = project_dir / "assets"
        assets.mkdir(parents=True, exist_ok=True)
        destination = assets / source.name
        shutil.copy2(source, destination)
        caption_text = _caption_label("figure", caption, rules)
        if caption_link:
            caption_text += r" \href{" + _escape(caption_link[0]) + "}{" + _escape(caption_link[1]) + "}"
        return (
            "\\begin{figure}["
            + str(policy.get("figure_float", "H"))
            + "]\n"
            + str(policy.get("figure_alignment", "\\centering"))
            + "\n\\includegraphics[width="
            + str(policy.get("figure_width", "0.85\\linewidth"))
            + "]{assets/"
            + destination.name
            + "}\n\\caption{"
            + caption_text
            + "}\n\\end{figure}"
        )
    rows = _read_table(content, upload)
    if not rows:
        raise ValueError("Provide a Markdown/CSV table or upload an Excel/CSV file.")
    width = max(len(row) for row in rows)
    column_spec = str(policy.get("table_columns", "l"))
    caption_text = _caption_label("table", caption, rules)
    if caption_link:
        caption_text += r" \href{" + _escape(caption_link[0]) + "}{" + _escape(caption_link[1]) + "}"
    lines = [
        "\\begin{table}[" + str(policy.get("table_float", "H")) + "]",
        str(policy.get("table_alignment", "\\centering")),
        "\\caption{" + caption_text + "}",
        "\\begin{tabular}{" + (column_spec * width) + "}",
        "\\toprule",
    ]
    for index, row in enumerate(rows):
        lines.append(" & ".join(_escape(value) for value in row + [""] * (width - len(row))) + r" \\")
        if index == 0:
            lines.append("\\midrule")
    lines.extend(["\\bottomrule", "\\end{tabular}", "\\end{table}"])
    return "\n".join(lines)


def replace_placeholder(tex: str, placeholder: str, block: str) -> str:
    return tex.replace(placeholder, block, 1)


def insert_block(tex: str, block: str, section: str | None, placement: str, anchor: str | None) -> str:
    section = section or ""
    anchor = anchor or ""
    marker = ""
    if anchor.strip():
        index = tex.find(anchor.strip())
        if index != -1:
            marker = anchor.strip()
            insert_at = index if placement == "Before anchor" else index + len(marker)
            return tex[:insert_at] + "\n\n" + block + "\n\n" + tex[insert_at:]
    if section.strip():
        match = re.search(r"\\section\{" + re.escape(section.strip()) + r"\}", tex, re.IGNORECASE)
        if match:
            if placement == "Section start":
                insert_at = match.end()
            else:
                following = re.search(r"\\section\{", tex[match.end():])
                insert_at = match.end() + following.start() if following else tex.find("\\end{document}")
            return tex[:insert_at] + "\n\n" + block + "\n\n" + tex[insert_at:]
    insert_at = tex.find("\\end{document}")
    return tex[:insert_at] + "\n\n" + block + "\n\n" + tex[insert_at:]
