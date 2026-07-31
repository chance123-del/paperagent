from __future__ import annotations

"""Parse user-supplied figure/table annotation workbooks without inventing text."""

from dataclasses import dataclass, field
from pathlib import Path
import re
import shutil
import zipfile


REQUIRED_SHEETS = {"Figures", "Tables", "Links"}
CAPTION_COLUMNS = {
    "asset_id", "caption_body", "caption_as_provided", "prefix_policy",
    "note", "source", "alt_text",
}
LINK_COLUMNS = {"asset_id", "url_or_doi", "link_text"}
PREFIX_POLICIES = {"template", "keep_as_provided", "review_required"}


@dataclass
class AnnotationData:
    figures: dict[str, str] = field(default_factory=dict)
    tables: dict[str, str] = field(default_factory=dict)
    links: dict[str, tuple[str, str]] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


def _key(value: object) -> str:
    normalized = str(value or "").strip().strip("[]").lower()
    normalized = normalized.replace("figure", "fig")
    return re.sub(r"[\s_.()\-]+", "", normalized)


def _safe_extract(archive: zipfile.ZipFile, destination: Path) -> None:
    destination_root = destination.resolve()
    for member in archive.infolist():
        target = (destination / member.filename).resolve()
        if not str(target).startswith(str(destination_root) + "\\") and target != destination_root:
            raise ValueError("Annotation ZIP contains an unsafe path.")
        if member.is_dir():
            continue
        target.parent.mkdir(parents=True, exist_ok=True)
        with archive.open(member) as source, target.open("wb") as output:
            shutil.copyfileobj(source, output)


def _workbook_path(upload: str | None, workspace: Path) -> Path | None:
    if not upload:
        return None
    source = Path(upload)
    if source.suffix.lower() == ".xlsx":
        return source
    if source.suffix.lower() != ".zip":
        raise ValueError("Annotation input must be annotations.xlsx or annotations.zip.")
    destination = workspace / "annotations_bundle"
    if destination.exists():
        shutil.rmtree(destination)
    destination.mkdir(parents=True)
    with zipfile.ZipFile(source, "r") as archive:
        _safe_extract(archive, destination)
    workbooks = list(destination.rglob("*.xlsx"))
    if len(workbooks) != 1:
        raise ValueError("Annotation ZIP must contain exactly one .xlsx workbook.")
    return workbooks[0]


def _row_values(sheet) -> list[dict[str, str]]:
    header = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = {header[index]: str(value or "").strip() for index, value in enumerate(row) if index < len(header)}
        if any(values.values()):
            rows.append(values)
    return rows


def _matching_prefix(value: str, asset_key: str, kind: str) -> bool:
    number = re.search(r"(\d+)", asset_key)
    if not number:
        return False
    label = "(?:fig(?:ure)?|图)" if kind == "figure" else "(?:table|表)"
    return bool(re.match(rf"^\s*{label}\.?\s*{number.group(1)}\b[\s:：.\-]*", value, re.IGNORECASE))


def _caption(row: dict[str, str], key: str, kind: str, warnings: list[str]) -> str:
    body = row.get("caption_body", "")
    provided = row.get("caption_as_provided", "")
    policy = row.get("prefix_policy", "").lower()
    if policy not in PREFIX_POLICIES:
        warnings.append(f"{kind} {key}: prefix_policy must be one of {sorted(PREFIX_POLICIES)}.")
        return ""
    if policy == "template":
        if body:
            return body
        if provided and _matching_prefix(provided, key, kind):
            return re.sub(r"^\s*(?:fig(?:ure)?|table|图|表)\.?\s*\d+\b[\s:：.\-]*", "", provided, flags=re.IGNORECASE)
        warnings.append(f"{kind} {key}: template policy needs caption_body, or a matching numbered prefix in caption_as_provided.")
        return ""
    if policy == "keep_as_provided" and provided:
        return provided
    warnings.append(f"{kind} {key}: caption requires review or is missing user-provided text.")
    return ""


def _normalise_url(value: str) -> str | None:
    value = value.strip()
    if value.lower().startswith("doi:"):
        value = value[4:].strip()
    if value.startswith("10."):
        return "https://doi.org/" + value
    if re.match(r"https?://[^\s]+$", value, re.IGNORECASE):
        return value
    return None


def load_annotations(upload: str | None, workspace: Path) -> AnnotationData:
    workbook_path = _workbook_path(upload, workspace)
    if not workbook_path:
        return AnnotationData()
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Annotation workbook support requires openpyxl. Install requirements.txt first.") from exc
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    missing = REQUIRED_SHEETS - set(workbook.sheetnames)
    if missing:
        raise ValueError(f"Annotation workbook is missing sheets: {', '.join(sorted(missing))}.")
    result = AnnotationData()
    for sheet_name, kind, target in (("Figures", "figure", result.figures), ("Tables", "table", result.tables)):
        sheet = workbook[sheet_name]
        header = {str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))}
        if not CAPTION_COLUMNS.issubset(header):
            raise ValueError(f"{sheet_name} is missing columns: {', '.join(sorted(CAPTION_COLUMNS - header))}.")
        for row in _row_values(sheet):
            key = _key(row.get("asset_id", ""))
            if not re.fullmatch(r"(?:fig|table)\d+", key):
                result.warnings.append(f"{sheet_name}: invalid asset_id '{row.get('asset_id', '')}'.")
                continue
            if (kind == "figure") != key.startswith("fig"):
                result.warnings.append(f"{sheet_name}: asset_id '{row.get('asset_id', '')}' has the wrong object type.")
                continue
            if key in target:
                result.warnings.append(f"{sheet_name}: duplicate asset_id '{row.get('asset_id', '')}'.")
                continue
            caption = _caption(row, key, kind, result.warnings)
            if caption:
                target[key] = caption
    links = workbook["Links"]
    header = {str(cell.value or "").strip() for cell in next(links.iter_rows(min_row=1, max_row=1))}
    if not LINK_COLUMNS.issubset(header):
        raise ValueError(f"Links is missing columns: {', '.join(sorted(LINK_COLUMNS - header))}.")
    for row in _row_values(links):
        key, url, text = _key(row.get("asset_id", "")), _normalise_url(row.get("url_or_doi", "")), row.get("link_text", "")
        if not re.fullmatch(r"(?:fig|table)\d+", key) or not url or not text:
            result.warnings.append(f"Links: invalid row for asset_id '{row.get('asset_id', '')}'; provide asset_id, DOI/URL, and link_text.")
        elif key in result.links:
            result.warnings.append(f"Links: duplicate asset_id '{row.get('asset_id', '')}'.")
        else:
            result.links[key] = (url, text)
    return result
